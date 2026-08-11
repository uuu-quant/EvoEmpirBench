#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import json
import glob
import pandas as pd
import numpy as np
from tabulate import tabulate
import matplotlib.pyplot as plt
from collections import defaultdict

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

from src.config.game_config import MODE_LEVEL1, MODE_LEVEL2, MODE_LEVEL3
from src.config.paths import RESULTS_DIR

def analyze_game1_results(results_dir, save_csv=True, save_xlsx=True):
    """
    分析Game1（寻路游戏）的评估结果
    
    Args:
        results_dir: 结果目录，通常是 .../outputs/results/{model}/game1
        save_csv: 是否保存为CSV文件
        save_xlsx: 是否保存为XLSX文件
    
    Returns:
        结果统计的DataFrame和详细数据的DataFrame
    """
    print(f"\n正在分析Game1结果: {results_dir}")
    
    # 查找评估结果文件
    result_files = glob.glob(os.path.join(results_dir, "evaluation_results_*.json"))
    
    if not result_files:
        print(f"未找到Game1结果文件: {results_dir}")
        return None, None
    
    # 使用最新的结果文件
    result_file = sorted(result_files)[-1]
    print(f"使用结果文件: {result_file}")
    
    try:
        with open(result_file, 'r') as f:
            results_data = json.load(f)
    except Exception as e:
        print(f"读取结果文件失败: {str(e)}")
        return None, None
    
    model_name = results_data.get("model", "unknown")
    
    # 初始化结果统计
    stats = []
    detailed_stats = []
    
    # 处理每个难度级别
    for mode in [MODE_LEVEL1, MODE_LEVEL2, MODE_LEVEL3]:
        if mode not in results_data.get("results", {}):
            continue
            
        mode_results = results_data["results"][mode]
        if not mode_results:
            continue
            
        # 计算统计指标
        total_maps = len(mode_results)
        success_count = sum(1 for r in mode_results if r.get("success", False))
        success_rate = success_count / total_maps if total_maps > 0 else 0
        
        avg_score = np.mean([r.get("custom_score", 0) for r in mode_results])
        avg_api_calls = np.mean([r.get("api_calls", 0) for r in mode_results])
        avg_exploration = np.mean([r.get("exploration_rate", 0) for r in mode_results])
        
        # 计算金币收集率
        total_coins = 5  # 假设每个地图有5个金币
        avg_coins_collected = np.mean([r.get("collected_coins", 0) for r in mode_results])
        coin_collection_rate = avg_coins_collected / total_coins
        
        # 计算平均剩余生命
        avg_lives = np.mean([r.get("lives_remaining", 0) for r in mode_results])
        
        # 对于Level 3，额外计算平均击杀怪物数和平均破坏障碍物数
        avg_killed_monsters = 0
        avg_destroyed_obstacles = 0
        if mode == MODE_LEVEL3:
            avg_killed_monsters = np.mean([r.get("killed_monsters", 0) for r in mode_results])
            avg_destroyed_obstacles = np.mean([r.get("destroyed_obstacles", 0) for r in mode_results])
            # 调整Level 3的得分，加上怪物击杀和障碍物破坏的奖励
            adjusted_score = avg_score + 500 * avg_killed_monsters + 500 * avg_destroyed_obstacles
        else:
            adjusted_score = avg_score
        
        # 记录统计结果
        stats_entry = {
            "模型": model_name,
            "难度级别": mode,
            "样本数": total_maps,
            "成功率": success_rate,
            "平均得分": adjusted_score,  # 使用调整后的得分
            "平均API调用数": avg_api_calls,
            "平均探索率": avg_exploration,
            "金币收集率": coin_collection_rate,
            "平均剩余生命": avg_lives
        }
        
        # 为Level 3添加额外的统计指标
        if mode == MODE_LEVEL3:
            stats_entry["平均击杀怪物数"] = avg_killed_monsters
            stats_entry["平均破坏障碍数"] = avg_destroyed_obstacles
            stats_entry["原始平均得分"] = avg_score  # 保存原始得分以供参考
        
        stats.append(stats_entry)
        
        # 收集每个level的详细数据
        for i, r in enumerate(mode_results):
            detailed_stats.append({
                "模型": model_name,
                "难度级别": mode,
                "地图编号": i + 1,
                "成功": r.get("success", False),
                "得分": r.get("custom_score", 0),
                "API调用数": r.get("api_calls", 0),
                "探索率": r.get("exploration_rate", 0),
                "收集金币数": r.get("collected_coins", 0),
                "剩余生命": r.get("lives_remaining", 0),
                "击杀怪物数": r.get("killed_monsters", 0) if mode == MODE_LEVEL3 else 0,
                "破坏障碍数": r.get("destroyed_obstacles", 0) if mode == MODE_LEVEL3 else 0
            })
    
    # 计算所有级别的综合统计
    if stats:
        # 计算所有级别的平均值
        all_modes_stats = {
            "模型": model_name,
            "难度级别": "所有级别",
            "样本数": sum(s["样本数"] for s in stats),
            "成功率": np.mean([s["成功率"] for s in stats]),
            "平均得分": np.mean([s["平均得分"] for s in stats]),
            "平均API调用数": np.mean([s["平均API调用数"] for s in stats]),
            "平均探索率": np.mean([s["平均探索率"] for s in stats]),
            "金币收集率": np.mean([s["金币收集率"] for s in stats]),
            "平均剩余生命": np.mean([s["平均剩余生命"] for s in stats])
        }
        
        # 添加Level 3特有的指标，如果存在
        level3_stats = next((s for s in stats if s["难度级别"] == MODE_LEVEL3), None)
        if level3_stats and "平均击杀怪物数" in level3_stats:
            all_modes_stats["平均击杀怪物数"] = level3_stats["平均击杀怪物数"]
            all_modes_stats["平均破坏障碍数"] = level3_stats["平均破坏障碍数"]
        
        stats.append(all_modes_stats)
    
    # 转换为DataFrame
    stats_df = pd.DataFrame(stats)
    detailed_df = pd.DataFrame(detailed_stats)
    
    # 格式化百分比
    format_cols = ["成功率", "平均探索率", "金币收集率"]
    for col in format_cols:
        if col in stats_df.columns:
            stats_df[col] = stats_df[col].apply(lambda x: f"{x:.2%}")
    
    if "探索率" in detailed_df.columns:
        detailed_df["探索率"] = detailed_df["探索率"].apply(lambda x: f"{x:.2%}")
    
    # 保存为CSV
    if save_csv:
        csv_path = os.path.join(results_dir, "game1_stats_summary.csv")
        stats_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        detailed_csv_path = os.path.join(results_dir, "game1_detailed_stats.csv")
        detailed_df.to_csv(detailed_csv_path, index=False, encoding='utf-8-sig')
        print(f"Game1统计结果已保存至: {csv_path} 和 {detailed_csv_path}")
    
    # 保存为XLSX
    if save_xlsx:
        xlsx_path = os.path.join(results_dir, "game1_stats_summary.xlsx")
        stats_df.to_excel(xlsx_path, index=False, engine='openpyxl')
        
        detailed_xlsx_path = os.path.join(results_dir, "game1_detailed_stats.xlsx")
        detailed_df.to_excel(detailed_xlsx_path, index=False, engine='openpyxl')
        print(f"Game1统计结果已保存至: {xlsx_path} 和 {detailed_xlsx_path}")
    
    # 打印结果
    print("\nGame1统计结果:")
    print(tabulate(stats_df, headers='keys', tablefmt='grid', showindex=False))
    
    return stats_df, detailed_df

def analyze_game2_results(results_dir, save_csv=True, save_xlsx=True):
    """
    分析Game2（消除游戏）的评估结果
    
    Args:
        results_dir: 结果目录，通常是 .../outputs/results/{model}/game2
        save_csv: 是否保存为CSV文件
        save_xlsx: 是否保存为XLSX文件
    
    Returns:
        结果统计的DataFrame和详细数据的DataFrame
    """
    print(f"\n正在分析Game2结果: {results_dir}")
    
    if not os.path.exists(results_dir):
        print(f"Game2结果目录不存在: {results_dir}")
        return None, None
    
    # 初始化结果收集
    results_data = defaultdict(list)
    # 从路径中提取模型名称
    model_name = os.path.basename(os.path.dirname(results_dir))
    if model_name == "game2":  # 如果直接给出了game2路径
        model_name = os.path.basename(os.path.dirname(os.path.dirname(results_dir)))
    
    # 难度级别
    difficulties = ["easy", "medium", "hard"]
    
    # 详细数据列表
    detailed_stats = []
    
    # 遍历所有难度和关卡
    for difficulty in difficulties:
        difficulty_dir = os.path.join(results_dir, difficulty)
        if not os.path.exists(difficulty_dir):
            continue
            
        for level_dir in sorted(glob.glob(os.path.join(difficulty_dir, "level*"))):
            agent_dir = os.path.join(level_dir, "agent")
            if not os.path.exists(agent_dir):
                continue
                
            # 寻找统计文件
            stats_files = glob.glob(os.path.join(agent_dir, "stats_*.json"))
            if not stats_files:
                continue
                
            # 使用最新的统计文件
            stats_file = sorted(stats_files)[-1]
            
            try:
                with open(stats_file, 'r') as f:
                    stats_data = json.load(f)
                    
                # 提取关卡号
                level_name = os.path.basename(level_dir)
                level_num = int(level_name.replace("level", ""))
                
                # 查找最大步数信息
                state_files = glob.glob(os.path.join(agent_dir, "game_state_*.json"))
                max_steps = 0
                if state_files:
                    try:
                        with open(sorted(state_files)[-1], 'r') as f:
                            state_data = json.load(f)
                            max_steps = state_data.get("max_steps", 0)
                    except Exception:
                        pass
                
                # 收集统计数据
                level_data = {
                    "关卡": level_num,
                    "得分": stats_data.get("total_score", 0),
                    "通关": stats_data.get("cleared", False),
                    "剩余步数": stats_data.get("steps_remaining", 0),
                    "最大步数": max_steps,
                    "剩余步数比例": stats_data.get("steps_remaining", 0) / max_steps if max_steps > 0 else 0,
                    "每步平均得分": stats_data.get("avg_score_per_step", 0),
                    "每步平均清除": stats_data.get("avg_clear_per_step", 0),
                    "API调用数": stats_data.get("api_response_count", 0),
                    "有效API响应数": stats_data.get("valid_api_response_count", 0),
                    "有效API响应比率": stats_data.get("valid_api_response_ratio", 0)
                }
                
                results_data[difficulty].append(level_data)
                
                # 添加到详细数据
                detailed_stats.append({
                    "模型": model_name,
                    "难度级别": difficulty,
                    **level_data
                })
                
            except Exception as e:
                print(f"处理文件失败 {stats_file}: {str(e)}")
    
    # 如果没有收集到数据
    if not results_data:
        print(f"未找到Game2统计数据: {results_dir}")
        return None, None
    
    # 汇总统计结果
    stats = []
    
    # 对每个难度级别计算统计值
    for difficulty in difficulties:
        if not results_data[difficulty]:
            continue
            
        difficulty_df = pd.DataFrame(results_data[difficulty])
        
        stats.append({
            "模型": model_name,
            "难度级别": difficulty,
            "样本数": len(difficulty_df),
            "平均得分": difficulty_df["得分"].mean(),
            "通关率": difficulty_df["通关"].mean(),
            "剩余步数/最大步数": difficulty_df["剩余步数比例"].mean(),
            "每步平均得分": difficulty_df["每步平均得分"].mean(),
            "每步平均清除": difficulty_df["每步平均清除"].mean(),
            "有效API响应比率": difficulty_df["有效API响应比率"].mean()
        })
    
    # 计算所有级别的综合统计
    all_results = []
    for difficulty in difficulties:
        all_results.extend(results_data[difficulty])
    
    if all_results:
        all_df = pd.DataFrame(all_results)
        stats.append({
            "模型": model_name,
            "难度级别": "所有级别",
            "样本数": len(all_df),
            "平均得分": all_df["得分"].mean(),
            "通关率": all_df["通关"].mean(),
            "剩余步数/最大步数": all_df["剩余步数比例"].mean(),
            "每步平均得分": all_df["每步平均得分"].mean(),
            "每步平均清除": all_df["每步平均清除"].mean(),
            "有效API响应比率": all_df["有效API响应比率"].mean()
        })
    
    # 转换为DataFrame
    stats_df = pd.DataFrame(stats)
    detailed_df = pd.DataFrame(detailed_stats)
    
    # 格式化百分比
    format_percent_cols = ["通关率", "剩余步数/最大步数", "有效API响应比率"]
    for col in format_percent_cols:
        if col in stats_df.columns:
            stats_df[col] = stats_df[col].apply(lambda x: f"{x:.2%}")
    
    if "剩余步数比例" in detailed_df.columns:
        detailed_df["剩余步数比例"] = detailed_df["剩余步数比例"].apply(lambda x: f"{x:.2%}")
    if "有效API响应比率" in detailed_df.columns:
        detailed_df["有效API响应比率"] = detailed_df["有效API响应比率"].apply(lambda x: f"{x:.2%}")
    
    # 保存为CSV
    if save_csv:
        csv_path = os.path.join(results_dir, "game2_stats_summary.csv")
        stats_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        detailed_csv_path = os.path.join(results_dir, "game2_detailed_stats.csv")
        detailed_df.to_csv(detailed_csv_path, index=False, encoding='utf-8-sig')
        print(f"Game2统计结果已保存至: {csv_path} 和 {detailed_csv_path}")
    
    # 保存为XLSX
    if save_xlsx:
        xlsx_path = os.path.join(results_dir, "game2_stats_summary.xlsx")
        stats_df.to_excel(xlsx_path, index=False, engine='openpyxl')
        
        detailed_xlsx_path = os.path.join(results_dir, "game2_detailed_stats.xlsx")
        detailed_df.to_excel(detailed_xlsx_path, index=False, engine='openpyxl')
        print(f"Game2统计结果已保存至: {xlsx_path} 和 {detailed_xlsx_path}")
    
    # 打印结果
    print("\nGame2统计结果:")
    print(tabulate(stats_df, headers='keys', tablefmt='grid', showindex=False))
    
    return stats_df, detailed_df

def compare_models(results_base_dir, models, save_csv=True, save_xlsx=True):
    """
    比较多个模型在Game1和Game2上的表现
    
    Args:
        results_base_dir: 结果基目录，通常是 .../outputs/results/
        models: 模型列表
        save_csv: 是否保存为CSV
        save_xlsx: 是否保存为XLSX
    """
    print(f"\n正在比较模型: {', '.join(models)}")
    
    # 收集所有模型的Game1结果
    game1_results = []
    game1_detailed = []
    for model in models:
        model_dir = os.path.join(results_base_dir, model, "game1")
        if os.path.exists(model_dir):
            stats_df, detailed_df = analyze_game1_results(model_dir, save_csv=False, save_xlsx=False)
            if stats_df is not None:
                game1_results.append(stats_df)
            if detailed_df is not None:
                game1_detailed.append(detailed_df)
    
    # 合并Game1结果
    if game1_results:
        game1_combined = pd.concat(game1_results, ignore_index=True)
        if save_csv:
            csv_path = os.path.join(results_base_dir, "game1_models_comparison.csv")
            game1_combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"Game1模型比较结果已保存至: {csv_path}")
        
        if save_xlsx:
            xlsx_path = os.path.join(results_base_dir, "game1_models_comparison.xlsx")
            game1_combined.to_excel(xlsx_path, index=False, engine='openpyxl')
            print(f"Game1模型比较结果已保存至: {xlsx_path}")
        
        # 创建Game1可视化
        create_game1_comparison_charts(game1_combined, results_base_dir)
    else:
        print("没有可用的Game1结果进行比较")
    
    if game1_detailed:
        game1_detailed_combined = pd.concat(game1_detailed, ignore_index=True)
        if save_csv:
            csv_path = os.path.join(results_base_dir, "game1_detailed_comparison.csv")
            game1_detailed_combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"Game1详细比较结果已保存至: {csv_path}")
        
        if save_xlsx:
            xlsx_path = os.path.join(results_base_dir, "game1_detailed_comparison.xlsx")
            game1_detailed_combined.to_excel(xlsx_path, index=False, engine='openpyxl')
            print(f"Game1详细比较结果已保存至: {xlsx_path}")
    
    # 收集所有模型的Game2结果
    game2_results = []
    game2_detailed = []
    for model in models:
        model_dir = os.path.join(results_base_dir, model, "game2")
        if os.path.exists(model_dir):
            stats_df, detailed_df = analyze_game2_results(model_dir, save_csv=False, save_xlsx=False)
            if stats_df is not None:
                game2_results.append(stats_df)
            if detailed_df is not None:
                game2_detailed.append(detailed_df)
    
    # 合并Game2结果
    if game2_results:
        game2_combined = pd.concat(game2_results, ignore_index=True)
        if save_csv:
            csv_path = os.path.join(results_base_dir, "game2_models_comparison.csv")
            game2_combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"Game2模型比较结果已保存至: {csv_path}")
        
        if save_xlsx:
            xlsx_path = os.path.join(results_base_dir, "game2_models_comparison.xlsx")
            game2_combined.to_excel(xlsx_path, index=False, engine='openpyxl')
            print(f"Game2模型比较结果已保存至: {xlsx_path}")
        
        # 创建Game2可视化
        create_game2_comparison_charts(game2_combined, results_base_dir)
    else:
        print("没有可用的Game2结果进行比较")
    
    if game2_detailed:
        game2_detailed_combined = pd.concat(game2_detailed, ignore_index=True)
        if save_csv:
            csv_path = os.path.join(results_base_dir, "game2_detailed_comparison.csv")
            game2_detailed_combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"Game2详细比较结果已保存至: {csv_path}")
        
        if save_xlsx:
            xlsx_path = os.path.join(results_base_dir, "game2_detailed_comparison.xlsx")
            game2_detailed_combined.to_excel(xlsx_path, index=False, engine='openpyxl')
            print(f"Game2详细比较结果已保存至: {xlsx_path}")

def create_game1_comparison_charts(stats_df, save_dir):
    """为Game1创建比较图表"""
    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    # 过滤出"所有级别"的数据进行比较
    all_levels_df = stats_df[stats_df["难度级别"] == "所有级别"].copy()
    
    # 将百分比字符串转换回浮点数
    for col in ["成功率", "平均探索率", "金币收集率"]:
        if col in all_levels_df.columns:
            all_levels_df[col] = all_levels_df[col].str.rstrip('%').astype('float') / 100
    
    # 1. 成功率比较
    plt.figure(figsize=(10, 6))
    plt.bar(all_levels_df["模型"], all_levels_df["成功率"])
    plt.title("不同模型在Game1中的成功率比较")
    plt.xlabel("模型")
    plt.ylabel("成功率")
    plt.ylim(0, 1)
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game1_success_rate_comparison.png"))
    plt.close()
    
    # 2. 平均得分比较
    plt.figure(figsize=(10, 6))
    plt.bar(all_levels_df["模型"], all_levels_df["平均得分"])
    plt.title("不同模型在Game1中的平均得分比较")
    plt.xlabel("模型")
    plt.ylabel("平均得分")
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game1_avg_score_comparison.png"))
    plt.close()
    
    # 3. 探索率和金币收集率比较
    plt.figure(figsize=(12, 6))
    width = 0.35
    x = np.arange(len(all_levels_df))
    
    plt.bar(x - width/2, all_levels_df["平均探索率"], width, label="平均探索率")
    plt.bar(x + width/2, all_levels_df["金币收集率"], width, label="金币收集率")
    
    plt.title("不同模型在Game1中的探索率和金币收集率比较")
    plt.xlabel("模型")
    plt.ylabel("比率")
    plt.xticks(x, all_levels_df["模型"])
    plt.ylim(0, 1)
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game1_exploration_coin_comparison.png"))
    plt.close()
    
    # 4. 不同难度级别下的成功率比较
    levels_df = stats_df[stats_df["难度级别"] != "所有级别"].copy()
    
    # 将百分比字符串转换回浮点数
    levels_df["成功率"] = levels_df["成功率"].str.rstrip('%').astype('float') / 100
    
    plt.figure(figsize=(12, 8))
    models = levels_df["模型"].unique()
    levels = [MODE_LEVEL1, MODE_LEVEL2, MODE_LEVEL3]
    
    x = np.arange(len(levels))
    width = 0.8 / len(models)
    
    for i, model in enumerate(models):
        model_data = levels_df[levels_df["模型"] == model]
        success_rates = []
        
        for level in levels:
            level_data = model_data[model_data["难度级别"] == level]
            if not level_data.empty:
                success_rates.append(level_data["成功率"].values[0])
            else:
                success_rates.append(0)
        
        plt.bar(x + (i - len(models)/2 + 0.5) * width, success_rates, width, label=model)
    
    plt.title("不同模型在各难度级别下的成功率比较")
    plt.xlabel("难度级别")
    plt.ylabel("成功率")
    plt.xticks(x, levels)
    plt.ylim(0, 1)
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game1_success_rate_by_level.png"))
    plt.close()

def create_game2_comparison_charts(stats_df, save_dir):
    """为Game2创建比较图表"""
    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    # 过滤出"所有级别"的数据进行比较
    all_levels_df = stats_df[stats_df["难度级别"] == "所有级别"].copy()
    
    # 将百分比字符串转换回浮点数
    for col in ["通关率", "剩余步数/最大步数", "有效API响应比率"]:
        if col in all_levels_df.columns:
            all_levels_df[col] = all_levels_df[col].str.rstrip('%').astype('float') / 100
    
    # 1. 通关率比较
    plt.figure(figsize=(10, 6))
    plt.bar(all_levels_df["模型"], all_levels_df["通关率"])
    plt.title("不同模型在Game2中的通关率比较")
    plt.xlabel("模型")
    plt.ylabel("通关率")
    plt.ylim(0, 1)
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game2_completion_rate_comparison.png"))
    plt.close()
    
    # 2. 平均得分比较
    plt.figure(figsize=(10, 6))
    plt.bar(all_levels_df["模型"], all_levels_df["平均得分"])
    plt.title("不同模型在Game2中的平均得分比较")
    plt.xlabel("模型")
    plt.ylabel("平均得分")
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game2_avg_score_comparison.png"))
    plt.close()
    
    # 3. 每步得分和每步清除比较
    plt.figure(figsize=(12, 6))
    width = 0.35
    x = np.arange(len(all_levels_df))
    
    plt.bar(x - width/2, all_levels_df["每步平均得分"], width, label="每步平均得分")
    plt.bar(x + width/2, all_levels_df["每步平均清除"] * 5, width, label="每步平均清除 (×5)")
    
    plt.title("不同模型在Game2中的每步平均得分和清除比较")
    plt.xlabel("模型")
    plt.ylabel("数值")
    plt.xticks(x, all_levels_df["模型"])
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game2_score_clear_per_step.png"))
    plt.close()
    
    # 4. 不同难度级别下的通关率比较
    levels_df = stats_df[stats_df["难度级别"] != "所有级别"].copy()
    
    # 将百分比字符串转换回浮点数
    levels_df["通关率"] = levels_df["通关率"].str.rstrip('%').astype('float') / 100
    
    plt.figure(figsize=(12, 8))
    models = levels_df["模型"].unique()
    difficulties = ["easy", "medium", "hard"]
    
    x = np.arange(len(difficulties))
    width = 0.8 / len(models)
    
    for i, model in enumerate(models):
        model_data = levels_df[levels_df["模型"] == model]
        completion_rates = []
        
        for difficulty in difficulties:
            difficulty_data = model_data[model_data["难度级别"] == difficulty]
            if not difficulty_data.empty:
                completion_rates.append(difficulty_data["通关率"].values[0])
            else:
                completion_rates.append(0)
        
        plt.bar(x + (i - len(models)/2 + 0.5) * width, completion_rates, width, label=model)
    
    plt.title("不同模型在各难度级别下的通关率比较")
    plt.xlabel("难度级别")
    plt.ylabel("通关率")
    plt.xticks(x, ["简单", "中等", "困难"])
    plt.ylim(0, 1)
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, "game2_completion_rate_by_level.png"))
    plt.close()

def generate_normalized_tables(results_base_dir, models, base_model="human"):
    """
    生成以特定模型为基准的标准化表格
    
    Args:
        results_base_dir: 结果基目录
        models: 要分析的模型列表
        base_model: 基准模型名称，默认为"human"
    
    Returns:
        标准化后的Game1表格
    """
    print(f"\n生成以{base_model}为基准的标准化表格...")
    
    # 确保基准模型在模型列表中
    if base_model not in models:
        print(f"警告: 基准模型{base_model}不在分析模型列表中！")
        return None
    
    # 收集Game1结果
    game1_results = []
    for model in models:
        model_dir = os.path.join(results_base_dir, model, "game1")
        if os.path.exists(model_dir):
            stats_df, _ = analyze_game1_results(model_dir, save_csv=False, save_xlsx=False)
            if stats_df is not None:
                game1_results.append(stats_df)
    
    if not game1_results:
        print("没有可用的Game1结果进行标准化")
        return None
    
    # 合并所有结果
    game1_combined = pd.concat(game1_results, ignore_index=True)
    
    # 将百分比字符串转换回浮点数
    for col in ["成功率", "平均探索率", "金币收集率"]:
        if col in game1_combined.columns:
            game1_combined[col] = game1_combined[col].str.rstrip('%').astype('float') / 100
    
    # 创建标准化表格
    normalized_data = []
    
    # 获取基准模型的数据
    base_data = game1_combined[game1_combined["模型"] == base_model].copy()
    
    if base_data.empty:
        print(f"错误: 无法找到基准模型{base_model}的数据！")
        return None
    
    # 难度级别映射，用于显示更友好的级别名称
    level_map = {
        "Level 1": "Level 1",
        "Level 2": "Level 2",
        "Level 3": "Level 3",
        "所有级别": "All Levels"
    }
    
    # 对每个难度级别进行标准化
    for level in ["Level 1", "Level 2", "Level 3", "所有级别"]:
        base_level_data = base_data[base_data["难度级别"] == level]
        
        if base_level_data.empty:
            continue
        
        # 获取基准值
        base_success_rate = base_level_data["成功率"].values[0]
        base_score = base_level_data["平均得分"].values[0]
        
        # 对每个模型进行标准化
        for model in models:
            model_level_data = game1_combined[(game1_combined["模型"] == model) & 
                                             (game1_combined["难度级别"] == level)]
            
            if model_level_data.empty:
                continue
            
            # 计算标准化值
            normalized_success_rate = model_level_data["成功率"].values[0] / base_success_rate if base_success_rate > 0 else 0
            normalized_score = model_level_data["平均得分"].values[0] / base_score if base_score > 0 else 0
            
            # 添加到结果中
            normalized_data.append({
                "模型": model,
                "难度级别": level_map[level],  # 使用映射后的级别名称
                "标准化成功率": normalized_success_rate,
                "标准化得分": normalized_score,
                "原始成功率": model_level_data["成功率"].values[0],
                "原始得分": model_level_data["平均得分"].values[0],
                "基准成功率": base_success_rate,
                "基准得分": base_score
            })
    
    # 转换为DataFrame
    normalized_df = pd.DataFrame(normalized_data)
    
    # 格式化百分比
    if not normalized_df.empty:
        # 将标准化值格式化为百分比
        normalized_df["标准化成功率"] = normalized_df["标准化成功率"].apply(lambda x: f"{x:.2%}")
        normalized_df["标准化得分"] = normalized_df["标准化得分"].apply(lambda x: f"{x:.2%}")
        
        # 将原始和基准成功率格式化为百分比
        normalized_df["原始成功率"] = normalized_df["原始成功率"].apply(lambda x: f"{x:.2%}")
        normalized_df["基准成功率"] = normalized_df["基准成功率"].apply(lambda x: f"{x:.2%}")
    
    # 打印表格
    if not normalized_df.empty:
        # 选择关键列进行显示
        display_df = normalized_df[["模型", "难度级别", "标准化成功率", "标准化得分"]].copy()
        
        print("\nGame1标准化指标（以human为基准）:")
        print(tabulate(display_df, headers='keys', tablefmt='grid', showindex=False))
        
        # 为Excel创建一个更优化的版本
        excel_df = normalized_df.copy()
        
        # 重新排序以便相同模型的不同级别分组在一起
        excel_df['level_order'] = excel_df['难度级别'].map({
            'Level 1': 1, 
            'Level 2': 2, 
            'Level 3': 3, 
            'All Levels': 4
        })
        excel_df = excel_df.sort_values(['模型', 'level_order']).drop('level_order', axis=1)
        
        # 保存到文件
        csv_path = os.path.join(results_base_dir, "game1_normalized.csv")
        excel_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        xlsx_path = os.path.join(results_base_dir, "game1_normalized.xlsx")
        
        # 使用ExcelWriter来应用样式
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            excel_df.to_excel(writer, index=False, sheet_name='标准化指标')
            
            # 获取工作表以应用格式
            workbook = writer.book
            worksheet = writer.sheets['标准化指标']
            
            # 设置列宽
            for i, col in enumerate(excel_df.columns):
                max_width = max(
                    excel_df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + i)].width = max_width
            
            # 应用表格样式
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            model_font = Font(bold=True)
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 为交替行应用不同的背景色
            light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # 为All Levels行应用突出显示
            all_level_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            # 应用行样式
            for i, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                level = excel_df.iloc[i-2]['难度级别']
                if level == 'All Levels':
                    for cell in row:
                        cell.fill = all_level_fill
                        cell.font = model_font
                elif i % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill
        
        print(f"标准化结果已保存至: {csv_path} 和 {xlsx_path}")
    
    return normalized_df

def generate_summary_tables(results_base_dir, models):
    """
    为Game1和Game2生成简化的摘要表格，只包含关键指标
    
    Args:
        results_base_dir: 结果基目录
        models: 要分析的模型列表
    
    Returns:
        游戏1和游戏2的汇总表格
    """
    print("\n正在生成摘要表格...")
    
    # 收集Game1结果
    game1_results = []
    
    # 级别映射，用于显示更明确的级别标识
    level_map = {
        "Level 1": "Level 1",
        "Level 2": "Level 2",
        "Level 3": "Level 3",
        "所有级别": "All Levels"
    }
    
    for model in models:
        model_dir = os.path.join(results_base_dir, model, "game1")
        if os.path.exists(model_dir):
            stats_df, _ = analyze_game1_results(model_dir, save_csv=False, save_xlsx=False)
            if stats_df is not None:
                # 为每个级别添加行，并标注明确的级别
                for level in ["Level 1", "Level 2", "Level 3", "所有级别"]:
                    level_data = stats_df[stats_df["难度级别"] == level].copy()
                    if not level_data.empty:
                        level_data["模型"] = model
                        level_data["难度级别"] = level_map[level]
                        game1_results.append(level_data)
    
    # 收集Game2结果
    game2_results = []
    
    # 难度映射
    difficulty_map = {
        "easy": "Easy",
        "medium": "Medium",
        "hard": "Hard",
        "所有级别": "All Levels"
    }
    
    for model in models:
        model_dir = os.path.join(results_base_dir, model, "game2")
        if os.path.exists(model_dir):
            stats_df, _ = analyze_game2_results(model_dir, save_csv=False, save_xlsx=False)
            if stats_df is not None:
                # 为每个难度添加行，并标注明确的难度
                for difficulty in ["easy", "medium", "hard", "所有级别"]:
                    diff_data = stats_df[stats_df["难度级别"] == difficulty].copy()
                    if not diff_data.empty:
                        diff_data["模型"] = model
                        diff_data["难度级别"] = difficulty_map[difficulty]
                        game2_results.append(diff_data)
    
    # 创建Game1摘要表格
    game1_summary = None
    if game1_results:
        game1_df = pd.concat(game1_results, ignore_index=True)
        
        # 将百分比字符串转换回浮点数用于排序
        for col in ["成功率", "平均探索率", "金币收集率"]:
            if col in game1_df.columns:
                game1_df[col] = game1_df[col].str.rstrip('%').astype('float') / 100
        
        # 选择需要的列并重命名
        selected_cols = ["模型", "难度级别", "成功率", "平均得分", "平均API调用数", 
                         "平均探索率", "金币收集率", "平均剩余生命"]
        
        # 添加Level 3特有的列
        if "平均击杀怪物数" in game1_df.columns:
            selected_cols.extend(["平均击杀怪物数", "平均破坏障碍数"])
        
        game1_summary = game1_df[selected_cols].copy()
        
        # 重新格式化百分比
        for col in ["成功率", "平均探索率", "金币收集率"]:
            if col in game1_summary.columns:
                game1_summary[col] = game1_summary[col].apply(lambda x: f"{x:.2%}")
                
        # 按模型和难度级别排序
        level_order = {
            'Level 1': 1,
            'Level 2': 2,
            'Level 3': 3,
            'All Levels': 4
        }
        game1_summary['level_order'] = game1_summary['难度级别'].map(level_order)
        game1_summary = game1_summary.sort_values(['模型', 'level_order']).drop('level_order', axis=1)
    
    # 创建Game2摘要表格
    game2_summary = None
    if game2_results:
        game2_df = pd.concat(game2_results, ignore_index=True)
        
        # 将百分比字符串转换回浮点数用于排序
        for col in ["通关率", "剩余步数/最大步数", "有效API响应比率"]:
            if col in game2_df.columns:
                game2_df[col] = game2_df[col].str.rstrip('%').astype('float') / 100
        
        # 选择需要的列
        selected_cols = ["模型", "难度级别", "平均得分", "通关率", "剩余步数/最大步数", 
                         "每步平均得分", "每步平均清除", "有效API响应比率"]
        
        game2_summary = game2_df[selected_cols].copy()
        
        # 重新格式化百分比
        for col in ["通关率", "剩余步数/最大步数", "有效API响应比率"]:
            if col in game2_summary.columns:
                game2_summary[col] = game2_summary[col].apply(lambda x: f"{x:.2%}")
        
        # 按模型和难度级别排序
        difficulty_order = {
            'Easy': 1,
            'Medium': 2,
            'Hard': 3,
            'All Levels': 4
        }
        game2_summary['difficulty_order'] = game2_summary['难度级别'].map(difficulty_order)
        game2_summary = game2_summary.sort_values(['模型', 'difficulty_order']).drop('difficulty_order', axis=1)
    
    # 打印表格
    if game1_summary is not None:
        print("\nGame1关键指标摘要:")
        print(tabulate(game1_summary, headers='keys', tablefmt='grid', showindex=False))
    
    if game2_summary is not None:
        print("\nGame2关键指标摘要:")
        print(tabulate(game2_summary, headers='keys', tablefmt='grid', showindex=False))
    
    # 保存为Excel格式，添加格式化
    if game1_summary is not None:
        xlsx_path = os.path.join(results_base_dir, "game1_summary.xlsx")
        
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            game1_summary.to_excel(writer, index=False, sheet_name='Game1指标')
            
            # 获取工作表以应用格式
            workbook = writer.book
            worksheet = writer.sheets['Game1指标']
            
            # 设置列宽
            for i, col in enumerate(game1_summary.columns):
                max_width = max(
                    game1_summary[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + i)].width = max_width
            
            # 应用表格样式
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            model_font = Font(bold=True)
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 为All Levels行应用突出显示
            all_level_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            # 应用行样式
            for i, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                level = game1_summary.iloc[i-2]['难度级别']
                if level == 'All Levels':
                    for cell in row:
                        cell.fill = all_level_fill
                        cell.font = model_font
        
        print(f"Game1摘要已保存至: {xlsx_path}")
    
    if game2_summary is not None:
        xlsx_path = os.path.join(results_base_dir, "game2_summary.xlsx")
        
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            game2_summary.to_excel(writer, index=False, sheet_name='Game2指标')
            
            # 获取工作表以应用格式
            workbook = writer.book
            worksheet = writer.sheets['Game2指标']
            
            # 设置列宽
            for i, col in enumerate(game2_summary.columns):
                max_width = max(
                    game2_summary[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + i)].width = max_width
            
            # 应用表格样式
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            model_font = Font(bold=True)
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 为All Levels行应用突出显示
            all_level_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            # 应用行样式
            for i, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                level = game2_summary.iloc[i-2]['难度级别']
                if level == 'All Levels':
                    for cell in row:
                        cell.fill = all_level_fill
                        cell.font = model_font
        
        print(f"Game2摘要已保存至: {xlsx_path}")
    
    return game1_summary, game2_summary

def main():
    """主函数"""
    import argparse
    parser = argparse.ArgumentParser(description="分析和比较不同模型在Game1和Game2上的表现")
    parser.add_argument("--models", nargs='+', help="要分析的模型列表")
    parser.add_argument("--results_dir", type=str, default=str(RESULTS_DIR),
                       help="结果文件的根目录")
    parser.add_argument("--game", type=str, choices=["game1", "game2", "both"], default="both",
                       help="要分析的游戏")
    parser.add_argument("--csv", action="store_true", help="保存CSV格式结果")
    parser.add_argument("--xlsx", action="store_true", help="保存XLSX格式结果")
    parser.add_argument("--summary", action="store_true", help="只生成摘要表格")
    parser.add_argument("--normalize", action="store_true", help="生成以human为基准的标准化表格")
    parser.add_argument("--base_model", type=str, default="human", help="标准化的基准模型")
    args = parser.parse_args()
    
    results_dir = args.results_dir
    save_csv = args.csv
    save_xlsx = args.xlsx or not args.csv  # 如果未指定CSV，默认保存XLSX
    
    # 如果未指定模型，查找结果目录中所有可用的模型
    if not args.models:
        models = [d for d in os.listdir(results_dir) 
                 if os.path.isdir(os.path.join(results_dir, d))]
    else:
        models = args.models
    
    print(f"将分析以下模型: {', '.join(models)}")
    
    # 如果需要标准化表格
    if args.normalize:
        base_model = args.base_model
        if base_model not in models:
            if os.path.isdir(os.path.join(results_dir, base_model)):
                models.append(base_model)
                print(f"已添加基准模型{base_model}到分析列表")
        
        generate_normalized_tables(results_dir, models, base_model)
        return
    
    # 如果只需要摘要表格
    if args.summary:
        game1_summary, game2_summary = generate_summary_tables(results_dir, models)
        
        # 保存摘要表格
        if game1_summary is not None:
            if save_csv:
                game1_summary.to_csv(os.path.join(results_dir, "game1_summary.csv"), 
                                   index=False, encoding='utf-8-sig')
            if save_xlsx:
                game1_summary.to_excel(os.path.join(results_dir, "game1_summary.xlsx"), 
                                     index=False, engine='openpyxl')
        
        if game2_summary is not None:
            if save_csv:
                game2_summary.to_csv(os.path.join(results_dir, "game2_summary.csv"), 
                                   index=False, encoding='utf-8-sig')
            if save_xlsx:
                game2_summary.to_excel(os.path.join(results_dir, "game2_summary.xlsx"), 
                                     index=False, engine='openpyxl')
        return
    
    # 根据指定的游戏类型分析结果
    if args.game in ["game1", "both"]:
        for model in models:
            model_game1_dir = os.path.join(results_dir, model, "game1")
            if os.path.exists(model_game1_dir):
                analyze_game1_results(model_game1_dir, save_csv=save_csv, save_xlsx=save_xlsx)
            else:
                print(f"模型 {model} 的Game1结果不存在")
    
    if args.game in ["game2", "both"]:
        for model in models:
            model_game2_dir = os.path.join(results_dir, model, "game2")
            if os.path.exists(model_game2_dir):
                analyze_game2_results(model_game2_dir, save_csv=save_csv, save_xlsx=save_xlsx)
            else:
                print(f"模型 {model} 的Game2结果不存在")
    
    # 比较不同模型的结果
    if len(models) > 1:
        compare_models(results_dir, models, save_csv=save_csv, save_xlsx=save_xlsx)

if __name__ == "__main__":
    main() 
